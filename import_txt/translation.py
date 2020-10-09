#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
import os
import re
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from typing import Optional, List

# This font can be gotten from https://github.com/adobe-fonts/source-han-mono/releases
EXCEL_FONT_NAME = 'Source Han Mono'


class Theme:
    font_name = EXCEL_FONT_NAME

    header_background_color_start = 'FCF3CF'
    header_background_color_end = 'FEF9E7'
    header_color = '333300'

    original_background_color_start = 'D6EAF8'
    original_background_color_end = 'EAFAF1'
    original_color = '000033'

    translation_background_color_start = 'D5F5E3'
    translation_background_color_end = 'EAFAF1'
    translation_color = '003300'

    comment_background_color_start = 'FADBD8'
    comment_background_color_end = 'FDEDEC'
    comment_color = '330000'

    def __init__(self):
        self.header_fill = openpyxl.styles.GradientFill(stop=(self.header_background_color_start,
                                                              self.header_background_color_end))
        self.header_font = openpyxl.styles.Font(name=self.font_name, bold=True, color=self.header_color)

        self.org_fill = openpyxl.styles.GradientFill(stop=(self.original_background_color_start,
                                                           self.original_background_color_end))
        self.org_font = openpyxl.styles.Font(name=self.font_name, color=self.original_color)

        self.trans_fill = openpyxl.styles.GradientFill(stop=(self.translation_background_color_start,
                                                             self.translation_background_color_end))
        self.trans_font = openpyxl.styles.Font(name=self.font_name, color=self.translation_color)

        self.comment_fill = openpyxl.styles.GradientFill(stop=(self.comment_background_color_start,
                                                               self.comment_background_color_end))
        self.comment_font = openpyxl.styles.Font(name=self.font_name, color=self.comment_color)

        self.font = openpyxl.styles.Font(name=self.font_name)

        self.border = Border(left=Side(style='hair'),
                             right=Side(style='hair'),
                             top=Side(style='hair'),
                             bottom=Side(style='hair'),
                             )


class Translation:
    def __init__(self,
                 name: str = None,
                 theme: Theme = Theme()):
        self.theme = theme
        if name is not None:
            self.load(name)

    def load(self, name: str):
        ext = os.path.splitext(name)[1].lower()
        if ext == '.xlsx':
            _df = pd.read_excel(name, engine='openpyxl')
            _df = _df.replace({'_x000D_': '\r'}, regex=True)
        elif ext == '.json':
            _df = pd.read_json(name)
        else:
            raise TypeError(f'not support type: "{ext}"')

        self.set_dataframe(_df)

    def save(self,
             name: str,
             index: str = 'English',
             drop_dup: bool = False):
        if drop_dup:
            self._df.drop_duplicates([index], inplace=True)
        ext = os.path.splitext(name)[1].lower()
        if ext == '.xlsx':
            self.save_excel(name, index)
        elif ext == '.json':
            self.save_json(name)
        else:
            raise TypeError(f'not support type: "{ext}"')

    def __set_excel_styles(self,
                           workbook: Workbook,
                           frezze_index: int):
        trans_rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual',
                                                         formula=['""'],
                                                         border=self.theme.border,
                                                         fill=self.theme.trans_fill,
                                                         font=self.theme.trans_font)
        comment_rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual',
                                                           formula=['""'],
                                                           border=self.theme.border,
                                                           fill=self.theme.comment_fill,
                                                           font=self.theme.comment_font)

        ws = workbook.active
        # the global style
        for i, row in enumerate(ws.iter_rows()):
            for cell in row:
                if cell.value is None:
                    cell.value = ''
                elif not isinstance(cell.value, str):
                    cell.value = str(cell.value)
                cell.number_format = '@'
                cell.data_type = 's'
                cell.quotePrefix = True

                if i == 0:
                    cell.font = self.theme.header_font
                    cell.fill = self.theme.header_fill
                else:
                    cell.font = self.theme.font

        # the comments style
        comment_letter = openpyxl.utils.get_column_letter(ws.max_column)
        if ws[comment_letter + '1'].value == 'Comments':
            ws.conditional_formatting.add(f'{comment_letter}2:{comment_letter}{ws.max_row}', comment_rule)
            cond_end = openpyxl.utils.get_column_letter(ws.max_column - 1)
        else:
            cond_end = openpyxl.utils.get_column_letter(ws.max_column)

        # the translation style
        cond_start = openpyxl.utils.get_column_letter(frezze_index + 1)
        ws.conditional_formatting.add(f'{cond_start}2:{cond_end}{ws.max_row}', trans_rule)

        # the original style
        start = openpyxl.utils.get_column_letter(frezze_index)
        index_cells = ws[f'{start}2:{start}{ws.max_row}']
        for cell in index_cells:
            cell[0].fill = self.theme.org_fill
            cell[0].font = self.theme.org_font
            cell[0].border = self.theme.border

    def save_excel(self,
                   name: str,
                   index: str = 'English'):
        frezze_index = self._df.columns.to_list().index(index) + 2
        self._df.to_excel(name, freeze_panes=(1, frezze_index), engine='openpyxl')
        wb = openpyxl.load_workbook(name)
        self.__set_excel_styles(wb, frezze_index)
        wb.save(name)

    def save_json(self, name: str):
        json_str = self._df.to_json(orient='records', force_ascii=False, indent=4)
        open(name, 'w', encoding='utf-8').write(json_str)

    def __process_dataframe(self):
        self._df.replace(float('nan'), '', inplace=True)
        self._df.drop(columns=filter(lambda x: 'Unnamed' in x or re.search(r'_\d', x), self._df.columns), inplace=True)

    def get_translation(self,
                        index: str = 'English',
                        fill_empty_with_org: bool = False,
                        empty_filter: bool = True):
        start_index = self._df.columns.to_list().index(index) + 1
        if empty_filter:
            rows = filter(lambda x: sum([len(y) for y in x[start_index:]]) > 0, self._df.itertuples(index=False))
            df = pd.DataFrame(rows)
        else:
            df = self._df.copy()
        df.set_index(index, drop=False, inplace=True)
        if fill_empty_with_org:
            for row in df.iterrows():
                for i, cell in enumerate(row[1]):
                    if len(cell) == 0:
                        row[1][i] = row[0]
        df.drop_duplicates([index], inplace=True)
        return df.to_dict('index')

    def set_dataframe(self,
                      df: pd.DataFrame,
                      add_comments: bool = True):
        if add_comments and 'Comments' not in df.columns:
            df['Comments'] = [''] * len(df)
        self._df = df
        self.__process_dataframe()

    def get_dataframe(self):
        return self._df

    def set_data(self,
                 data: List,
                 columns: Optional[List] = None):
        if columns is None:
            columns = data[0].keys()
        self.set_dataframe(pd.DataFrame(data, columns=columns, dtype=str))

    def get_percent(self, target_index: str):
        count = len(self._df[self._df[target_index] != ''])
        return count / len(self._df.index) * 100

    def check_variables(self,
                        regex: str,
                        org_index: str,
                        trans_index: str,
                        ordered: bool = True):
        resuls = []
        for i, row in self._df.iterrows():
            org_vars = re.findall(regex, row[org_index])
            trans_vars = re.findall(regex, row[trans_index])
            if len(org_vars) != len(trans_vars):
                resuls.append([i, org_vars, trans_vars])
            else:
                if not ordered:
                    unordered_org_vars = sorted(org_vars)
                    unordered_trans_vars = sorted(trans_vars)
                if unordered_org_vars != unordered_trans_vars:
                    resuls.append([i, org_vars, trans_vars])
        return resuls

    def check_size(self,
                   org_index: str,
                   trans_index: str,
                   encoding: str):
        result = []
        for i, row in self._df.iterrows():
            org = row[org_index]
            trans = row[trans_index]
            org_size = len(org.encode(encoding))
            trans_size = len(trans.encode(encoding))
            if org_size < trans_size:
                result.append([i, org_size, trans_size, org, trans])
        return result

    def __iter__(self):
        for _, row in self._df.iterrows():
            yield row


def try_to_get_translation(name: str):
    if os.path.exists(name):
        return Translation(name).get_translation()
    else:
        return {}
